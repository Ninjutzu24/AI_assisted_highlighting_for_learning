import json
import os
from collections import defaultdict, Counter
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference


RESULTS_FILE = "results.json"
SUS_RESULTS_FILE = "sus_nasa_results.json"
CHATBOT_LOG_FILE = "chatbot_logs.json"
OUTPUT_FILE = "results_analysis.xlsx"


RAW_HEADERS = [
    "participantId",
    "mode",
    "articleId",
    "articleIndex",
    "correct",
    "total",
    "score",
    "startedAt",
    "quizOpenedAt",
    "submittedAt",
    "readingTimeSeconds",
    "totalTaskTimeSeconds",
    "timeLimitSeconds",
    "timeLimitReached",
    "manualHighlightCount",
    "manualHighlightRemoveCount",
    "aiHighlightCount",
    "interactiveCardsShown",
    "chatbotQuestionCount",
    "quizQuestionCount",
    "quizAnsweredCount",
]


# IMPORTANT:
# These keys must match the keys saved inside sus_nasa_results.json.
# The current application saves the standardized SUS + NASA-TLX-style form
# using keys such as sus_frequent_use, sus_easy_to_use, nasa_mental_demand, etc.
FEEDBACK_RESPONSE_KEYS = [
    "sus_frequent_use",
    "sus_unnecessarily_complex",
    "sus_easy_to_use",
    "sus_need_technical_support",
    "sus_well_integrated",
    "sus_inconsistent",
    "sus_quick_to_learn",
    "sus_cumbersome",
    "sus_confident",
    "sus_need_to_learn",
    "nasa_mental_demand",
    "nasa_physical_demand",
    "nasa_temporal_demand",
    "nasa_performance",
    "nasa_effort",
    "nasa_frustration",
    "nasa_pair_1",
    "nasa_pair_2",
    "nasa_pair_3",
    "nasa_pair_4",
    "nasa_pair_5",
    "nasa_pair_6",
    "nasa_pair_7",
    "nasa_pair_8",
    "nasa_pair_9",
    "nasa_pair_10",
    "nasa_pair_11",
    "nasa_pair_12",
    "nasa_pair_13",
    "nasa_pair_14",
    "nasa_pair_15",
]


FEEDBACK_RAW_HEADERS = [
    "participantId",
    "taskNumber",
    "articleId",
    "articleIndex",
    "mode",
    "modeLabel",
    *FEEDBACK_RESPONSE_KEYS,
    "susScore",
    "rawNasaWorkload",
    "weightedNasaWorkload",
    "comment",
    "submittedAt",
]

CHATBOT_LOG_HEADERS = [
    "participantId",
    "articleId",
    "articleIndex",
    "mode",
    "question",
    "answer",
    "submittedAt",
]


SUS_POSITIVE_KEYS = [
    "sus_frequent_use",       # SUS 1
    "sus_easy_to_use",        # SUS 3
    "sus_well_integrated",    # SUS 5
    "sus_quick_to_learn",     # SUS 7
    "sus_confident",          # SUS 9
]


SUS_NEGATIVE_KEYS = [
    "sus_unnecessarily_complex",    # SUS 2
    "sus_need_technical_support",   # SUS 4
    "sus_inconsistent",             # SUS 6
    "sus_cumbersome",               # SUS 8
    "sus_need_to_learn",            # SUS 10
]


NASA_RATING_KEYS = [
    "nasa_mental_demand",
    "nasa_physical_demand",
    "nasa_temporal_demand",
    "nasa_performance",
    "nasa_effort",
    "nasa_frustration",
]


NASA_LABEL_TO_KEY = {
    "Mental Demand": "nasa_mental_demand",
    "Physical Demand": "nasa_physical_demand",
    "Temporal Demand": "nasa_temporal_demand",
    "Performance": "nasa_performance",
    "Effort": "nasa_effort",
    "Frustration": "nasa_frustration",
}


def safe_number(value, default=0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (ValueError, TypeError):
        return default


def avg(rows, key):
    values = [
        safe_number(row.get(key))
        for row in rows
        if row.get(key) is not None
    ]

    if not values:
        return 0

    return round(mean(values), 2)


def avg_values(values):
    clean = [safe_number(v) for v in values if v is not None and v != ""]

    if not clean:
        return 0

    return round(mean(clean), 2)


def load_json_list(file_path):
    if not os.path.exists(file_path):
        return []

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return []

    if not isinstance(data, list):
        return []

    return data


def load_results():
    return load_json_list(RESULTS_FILE)


def load_feedback_results():
    return load_json_list(SUS_RESULTS_FILE)


def load_chatbot_logs():
    return load_json_list(CHATBOT_LOG_FILE)


def get_mode_label(mode, fallback=None):
    if fallback:
        return fallback

    if mode == "static":
        return "Manual Highlighting"

    if mode == "interactive":
        return "AI Highlighting and Main Ideas Support"

    if mode == "chatbot":
        return "AI Chatbot Support"

    return mode or "unknown"


def get_responses(row):
    responses = row.get("responses", {})

    if not isinstance(responses, dict):
        return {}

    return responses


def get_response_value(row, key):
    return get_responses(row).get(key)


def avg_response_key(rows, key):
    values = [
        get_response_value(row, key)
        for row in rows
        if get_response_value(row, key) is not None
    ]

    return avg_values(values)


def calculate_sus_score(row):
    """
    Standard SUS scoring:
    - Positive items 1,3,5,7,9: score contribution = answer - 1
    - Negative items 2,4,6,8,10: score contribution = 5 - answer
    - Total contribution is multiplied by 2.5, giving a score from 0 to 100.
    """
    responses = get_responses(row)

    required_keys = SUS_POSITIVE_KEYS + SUS_NEGATIVE_KEYS
    if any(responses.get(key) is None for key in required_keys):
        return None

    total = 0

    for key in SUS_POSITIVE_KEYS:
        total += safe_number(responses.get(key)) - 1

    for key in SUS_NEGATIVE_KEYS:
        total += 5 - safe_number(responses.get(key))

    return round(total * 2.5, 2)


def calculate_raw_nasa_workload(row):
    responses = get_responses(row)

    values = [
        responses.get(key)
        for key in NASA_RATING_KEYS
        if responses.get(key) is not None
    ]

    if not values:
        return None

    return round(mean([safe_number(v) for v in values]), 2)


def calculate_weighted_nasa_workload(row):
    """
    Weighted NASA-TLX-style score using the 15 pairwise choices.
    If pairwise choices are missing, this falls back to raw NASA workload.
    The result remains on a 0-20 scale.
    """
    responses = get_responses(row)

    pair_values = [
        responses.get(f"nasa_pair_{i}")
        for i in range(1, 16)
        if responses.get(f"nasa_pair_{i}")
    ]

    if not pair_values:
        return calculate_raw_nasa_workload(row)

    weights = Counter(pair_values)

    weighted_sum = 0
    total_weight = 0

    for label, key in NASA_LABEL_TO_KEY.items():
        rating = responses.get(key)

        if rating is None:
            continue

        weight = weights.get(label, 0)
        weighted_sum += safe_number(rating) * weight
        total_weight += weight

    if total_weight == 0:
        return calculate_raw_nasa_workload(row)

    return round(weighted_sum / total_weight, 2)


def avg_calculated(rows, calc_function):
    values = []

    for row in rows:
        value = calc_function(row)
        if value is not None:
            values.append(value)

    return avg_values(values)


def style_sheet(ws):
    header_fill = PatternFill(
        start_color="1D4ED8",
        end_color="1D4ED8",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 38)


def clean_chart(chart, width=17, height=8, legend_position="b", show_legend=True):
    chart.height = height
    chart.width = width
    chart.x_axis.title = None
    chart.y_axis.title = None

    if show_legend:
        if chart.legend:
            chart.legend.position = legend_position
            chart.legend.overlay = False
    else:
        chart.legend = None


def add_raw_sheet(wb, results):
    ws = wb.create_sheet("Raw Results")

    ws.append(RAW_HEADERS)

    for row in results:
        ws.append([
            row.get(header)
            for header in RAW_HEADERS
        ])

    style_sheet(ws)


def add_feedback_raw_sheet(wb, feedback_results):
    ws = wb.create_sheet("Feedback Raw")

    ws.append(FEEDBACK_RAW_HEADERS)

    for row in feedback_results:
        responses = get_responses(row)

        mode = row.get("mode")
        mode_label = get_mode_label(
            mode,
            row.get("modeLabel")
        )

        ws.append([
            row.get("participantId"),
            row.get("taskNumber"),
            row.get("articleId"),
            row.get("articleIndex"),
            mode,
            mode_label,
            *[
                responses.get(key)
                for key in FEEDBACK_RESPONSE_KEYS
            ],
            calculate_sus_score(row),
            calculate_raw_nasa_workload(row),
            calculate_weighted_nasa_workload(row),
            row.get("comment"),
            row.get("submittedAt"),
        ])

    style_sheet(ws)




def add_chatbot_logs_sheet(wb, chatbot_logs):
    ws = wb.create_sheet("Chatbot Logs")

    ws.append(CHATBOT_LOG_HEADERS)

    for row in chatbot_logs:
        ws.append([
            row.get("participantId"),
            row.get("articleId"),
            row.get("articleIndex"),
            get_mode_label(row.get("mode")),
            row.get("question"),
            row.get("answer"),
            row.get("submittedAt"),
        ])

    style_sheet(ws)

    # Make question and answer columns easier to read.
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 80

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_summary_by_key(results, group_key):
    grouped = defaultdict(list)

    for row in results:
        key = row.get(group_key, "unknown")
        grouped[str(key)].append(row)

    summary = []

    for key, rows in sorted(grouped.items()):
        summary.append({
            group_key: key,
            "n": len(rows),
            "avgScore": avg(rows, "score"),
            "avgReadingTimeSeconds": avg(rows, "readingTimeSeconds"),
            "avgTotalTaskTimeSeconds": avg(rows, "totalTaskTimeSeconds"),
            "avgManualHighlights": avg(rows, "manualHighlightCount"),
            "avgAIHighlights": avg(rows, "aiHighlightCount"),
            "avgChatbotQuestions": avg(rows, "chatbotQuestionCount"),
            "avgQuizAnswered": avg(rows, "quizAnsweredCount"),
            "timeLimitReachedCount": sum(
                1 for row in rows
                if row.get("timeLimitReached") is True
            )
        })

    return summary


def build_feedback_summary_by_key(feedback_results, group_key):
    grouped = defaultdict(list)

    for row in feedback_results:
        if group_key == "mode":
            key = get_mode_label(
                row.get("mode"),
                row.get("modeLabel")
            )
        else:
            key = row.get(group_key, "unknown")

        grouped[str(key)].append(row)

    summary = []

    for key, rows in sorted(grouped.items()):
        summary.append({
            group_key: key,
            "n": len(rows),
            "avgSUSScore": avg_calculated(rows, calculate_sus_score),
            "avgRawNASAWorkload": avg_calculated(rows, calculate_raw_nasa_workload),
            "avgWeightedNASAWorkload": avg_calculated(rows, calculate_weighted_nasa_workload),
            "avgMentalDemand": avg_response_key(rows, "nasa_mental_demand"),
            "avgPhysicalDemand": avg_response_key(rows, "nasa_physical_demand"),
            "avgTemporalDemand": avg_response_key(rows, "nasa_temporal_demand"),
            "avgPerformance": avg_response_key(rows, "nasa_performance"),
            "avgEffort": avg_response_key(rows, "nasa_effort"),
            "avgFrustration": avg_response_key(rows, "nasa_frustration"),
            "avgSUSFrequentUse": avg_response_key(rows, "sus_frequent_use"),
            "avgSUSEasyToUse": avg_response_key(rows, "sus_easy_to_use"),
            "avgSUSConfident": avg_response_key(rows, "sus_confident"),
        })

    return summary


def add_summary_sheet(wb, title, group_key, summary):
    ws = wb.create_sheet(title)

    headers = [
        group_key,
        "n",
        "avgScore",
        "avgReadingTimeSeconds",
        "avgTotalTaskTimeSeconds",
        "avgManualHighlights",
        "avgAIHighlights",
        "avgChatbotQuestions",
        "avgQuizAnswered",
        "timeLimitReachedCount",
    ]

    ws.append(headers)

    for row in summary:
        ws.append([
            row.get(header)
            for header in headers
        ])

    style_sheet(ws)


def add_feedback_summary_sheet(wb, title, group_key, summary):
    ws = wb.create_sheet(title)

    headers = [
        group_key,
        "n",
        "avgSUSScore",
        "avgRawNASAWorkload",
        "avgWeightedNASAWorkload",
        "avgMentalDemand",
        "avgPhysicalDemand",
        "avgTemporalDemand",
        "avgPerformance",
        "avgEffort",
        "avgFrustration",
        "avgSUSFrequentUse",
        "avgSUSEasyToUse",
        "avgSUSConfident",
    ]

    ws.append(headers)

    for row in summary:
        ws.append([
            row.get(header)
            for header in headers
        ])

    style_sheet(ws)


def add_quiz_charts(ws, wb, start_cell="A3"):
    if "Summary by Mode" not in wb.sheetnames:
        ws["A3"] = "No quiz results available yet."
        return

    summary_mode = wb["Summary by Mode"]
    max_row = summary_mode.max_row

    if max_row < 2:
        ws["A3"] = "Not enough quiz data for charts."
        return

    categories = Reference(
        summary_mode,
        min_col=1,
        min_row=2,
        max_row=max_row
    )

    chart_score = BarChart()
    chart_score.type = "col"
    chart_score.title = "Average comprehension score by mode"

    data = Reference(
        summary_mode,
        min_col=3,
        min_row=1,
        max_row=max_row
    )

    chart_score.add_data(data, titles_from_data=True)
    chart_score.set_categories(categories)
    clean_chart(chart_score, width=16, height=8, show_legend=False)

    ws.add_chart(chart_score, "A3")

    chart_time = BarChart()
    chart_time.type = "col"
    chart_time.title = "Average total task time by mode"

    data = Reference(
        summary_mode,
        min_col=5,
        min_row=1,
        max_row=max_row
    )

    chart_time.add_data(data, titles_from_data=True)
    chart_time.set_categories(categories)
    clean_chart(chart_time, width=16, height=8, show_legend=False)

    ws.add_chart(chart_time, "A20")

    chart_interactions = BarChart()
    chart_interactions.type = "col"
    chart_interactions.title = "Interaction metrics by mode"

    data = Reference(
        summary_mode,
        min_col=6,
        max_col=8,
        min_row=1,
        max_row=max_row
    )

    chart_interactions.add_data(data, titles_from_data=True)
    chart_interactions.set_categories(categories)
    clean_chart(chart_interactions, width=18, height=8, legend_position="b")

    ws.add_chart(chart_interactions, "A37")


def add_feedback_charts(ws, wb):
    if "Feedback by Mode" not in wb.sheetnames:
        ws["J3"] = "No feedback results available yet."
        return

    feedback_mode = wb["Feedback by Mode"]
    max_row = feedback_mode.max_row

    if max_row < 2:
        ws["J3"] = "Not enough feedback data for charts."
        return

    categories = Reference(
        feedback_mode,
        min_col=1,
        min_row=2,
        max_row=max_row
    )

    chart_feedback = BarChart()
    chart_feedback.type = "col"
    chart_feedback.title = "SUS and NASA workload by mode"

    data = Reference(
        feedback_mode,
        min_col=3,
        max_col=5,
        min_row=1,
        max_row=max_row
    )

    chart_feedback.add_data(data, titles_from_data=True)
    chart_feedback.set_categories(categories)
    clean_chart(chart_feedback, width=18, height=8, legend_position="b")

    ws.add_chart(chart_feedback, "J3")

    chart_nasa = BarChart()
    chart_nasa.type = "col"
    chart_nasa.title = "NASA-TLX dimensions by mode"

    data = Reference(
        feedback_mode,
        min_col=6,
        max_col=11,
        min_row=1,
        max_row=max_row
    )

    chart_nasa.add_data(data, titles_from_data=True)
    chart_nasa.set_categories(categories)
    clean_chart(chart_nasa, width=18, height=8, legend_position="b")

    ws.add_chart(chart_nasa, "J20")

    chart_sus = BarChart()
    chart_sus.type = "col"
    chart_sus.title = "Selected SUS item averages by mode"

    data = Reference(
        feedback_mode,
        min_col=12,
        max_col=14,
        min_row=1,
        max_row=max_row
    )

    chart_sus.add_data(data, titles_from_data=True)
    chart_sus.set_categories(categories)
    clean_chart(chart_sus, width=18, height=8, legend_position="b")

    ws.add_chart(chart_sus, "J37")


def add_charts_sheet(wb):
    ws = wb.create_sheet("Charts")

    ws["A1"] = "Automatic Charts"
    ws["A1"].font = Font(size=18, bold=True)

    ws["A2"] = "Comprehension quiz and interaction metrics"
    ws["A2"].font = Font(size=13, bold=True)

    ws["J2"] = "Feedback questionnaire metrics"
    ws["J2"].font = Font(size=13, bold=True)

    add_quiz_charts(ws, wb)
    add_feedback_charts(ws, wb)


def export_results_to_excel():
    results = load_results()
    feedback_results = load_feedback_results()
    chatbot_logs = load_chatbot_logs()

    if not results and not feedback_results and not chatbot_logs:
        print("[Excel Export] No results, feedback data, or chatbot logs yet.")
        return False

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    if results:
        add_raw_sheet(wb, results)

        summary_by_mode = build_summary_by_key(
            results,
            "mode"
        )

        add_summary_sheet(
            wb,
            "Summary by Mode",
            "mode",
            summary_by_mode
        )

        summary_by_article = build_summary_by_key(
            results,
            "articleIndex"
        )

        add_summary_sheet(
            wb,
            "Summary by Article",
            "articleIndex",
            summary_by_article
        )

        summary_by_participant = build_summary_by_key(
            results,
            "participantId"
        )

        add_summary_sheet(
            wb,
            "Summary by Participant",
            "participantId",
            summary_by_participant
        )

    if feedback_results:
        add_feedback_raw_sheet(wb, feedback_results)

        feedback_by_mode = build_feedback_summary_by_key(
            feedback_results,
            "mode"
        )

        add_feedback_summary_sheet(
            wb,
            "Feedback by Mode",
            "mode",
            feedback_by_mode
        )

        feedback_by_article = build_feedback_summary_by_key(
            feedback_results,
            "articleIndex"
        )

        add_feedback_summary_sheet(
            wb,
            "Feedback by Article",
            "articleIndex",
            feedback_by_article
        )

        feedback_by_participant = build_feedback_summary_by_key(
            feedback_results,
            "participantId"
        )

        add_feedback_summary_sheet(
            wb,
            "Feedback by Participant",
            "participantId",
            feedback_by_participant
        )

    if chatbot_logs:
        add_chatbot_logs_sheet(wb, chatbot_logs)

    add_charts_sheet(wb)

    wb.save(OUTPUT_FILE)

    print(f"[Excel Export] Updated {OUTPUT_FILE}")
    return True


def main():
    export_results_to_excel()


if __name__ == "__main__":
    main()


